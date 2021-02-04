import requests
import os

import mysql.connector

from datetime import datetime
from bs4 import BeautifulSoup
from openpyxl import Workbook
from dotenv import load_dotenv

from constants import URL


def scrape():
    '''Function to scrape data from Amazon search results page and put it into an xls sheet'''
    HEADERS = ({'User-Agent':
                'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/44.0.2403.157 Safari/537.36',
                'Accept-Language': 'en-US, en;q=0.5'})

    # HTTP Request
    webpage = requests.get(URL, headers=HEADERS)
    # Soup Object containing all data
    soup = BeautifulSoup(webpage.content, "html.parser")
    # Calling search function thaat scrapes items and returns them in a list
    items = search(soup)
    # Calling the function to extract items into an xls sheet
    extract_in_xls(items)
    # Calling the function to extract items into database
    # extract_in_db(items)

def search(soup):
    '''Function to get list of items on search results page on Amazon'''
    # A list to store details of the items
    items = []
    for item in soup.find_all('div', attrs={'class': 's-result-item'}):
        # Getting various fields using respective html tag classes
        name = item.find('span', attrs={'class': 'a-size-medium a-color-base a-text-normal'})
        rating = item.find('span', attrs={'class': 'a-icon-alt'})
        rating_count = item.find('span', attrs={'class': 'a-size-base'})
        price_symbol = item.find('span', attrs={'class': 'a-price-symbol'})
        price = item.find('span', attrs={'class': 'a-price-whole'})
        if None in (name, rating, rating_count, price_symbol, price):
            # Condition to check if any field is None.
            # If any of the field is none, then the item is not a search result as Amazon uses same classes for other elements as well.
            continue
        else:
            # Extracting text from bs4 tag elements and appending a list of the fields into the items list
            items.append(
                [ name.get_text(),
                float(rating.get_text().split()[0]),
                int(rating_count.get_text().replace(",", "")),
                price_symbol.get_text(),
                int(price.get_text().replace(",", "")) ]
            )
    return items

def extract_in_xls(items):
    ''' A function to extract items in the items list into an xls sheet'''
    wb = Workbook()
    ws = wb.active
    # Assigning names to Column
    ws['A1'] = 'Title'
    ws['B1'] = 'Ratings'
    ws['C1'] = 'Rated By'
    ws['D1'] = 'Price Symbol'
    ws['E1'] = 'Price'
    # Appending rows into xls sheet
    for item in items:
        ws.append(item)
    # Saving the worksheet into a local xls sheet
    wb.save("result.xlsx")

def extract_in_db(items):
    # Loading environment variables from .env file
    load_dotenv()
    # Connecting to the database using credentials retreived from .env file
    mydb = mysql.connector.connect(
      host = os.environ.get("mysql_host"),
      user = os.environ.get("mysql_user"),
      password = os.environ.get("mysql_password"),
      database = os.environ.get("mysql_database"),
    )
    mycursor = mydb.cursor()
    # Iterating through all the scraped items
    for item in items:
        sql_fetch_item = "Select id from items where name=%s"
        val_fetch_item = (item[0], )
        # Command to get the item's id from the items table if it already exists in it.
        mycursor.execute(sql_fetch_item, val_fetch_item)
        item_curr = mycursor.fetchall()
        if not item_curr:
            # Condition to check if the item doesn't already exist in the table
            sql_insert_item = "Insert into items (name) values (%s)"
            val_insert_item = (item[0], )
            # Command to insert the item into the table
            mycursor.execute(sql_insert_item, val_insert_item)
            # Command to fetch the id of inserted item from the table
            mycursor.execute(sql_fetch_item, val_fetch_item)
            item_curr = mycursor.fetchall()
        item_id = item_curr[0][0]
        # Retreiving current date-time and formatting it according to MySQL
        now = datetime.now()
        formatted_date = now.strftime('%Y-%m-%d %H:%M:%S')
        sql_insert_price = "Insert into prices (price, itemid, record_date) values (%s, %s, %s)"
        val_insert_price = (item[4], item_id, formatted_date)
        # Inserting the price of the item into the price table
        mycursor.execute(sql_insert_price, val_insert_price)
    mydb.commit()

def print_result(items):
    counter=1
    for item in items:
        print(counter)
        for detail in item:
            print(detail)
        counter+=1

if __name__ == '__main__':
    '''main function'''
    scrape()
